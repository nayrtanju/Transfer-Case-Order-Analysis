import os
import re
import tempfile
import traceback
from io import BytesIO
from typing import Dict, Mapping, Optional, Tuple

import altair as alt
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st

try:
    import plotly.graph_objects as go
except ImportError:
    go = None

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill


# =============================================================================
# PAGE CONFIGURATION
# =============================================================================

st.set_page_config(
    page_title="NVH Analysis Suite",
    page_icon="⚙️",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# =============================================================================
# UX SETTINGS, LANGUAGE AND THEME
# =============================================================================

LANGUAGES = {
    "English": {
        "run": "Run Analysis",
        "pdf": "Download PDF Report",
        "startup": "Engineering analysis environment initialized",
    },
    "Türkçe": {
        "run": "Analizi Başlat",
        "pdf": "PDF Raporunu İndir",
        "startup": "Mühendislik analiz ortamı hazırlandı",
    },
}

with st.sidebar:
    language_name = st.selectbox(
        "Language / Dil",
        list(LANGUAGES.keys()),
        index=0,
        key="language_name",
    )

    theme_mode = st.selectbox(
        "Appearance",
        ["Corporate Light", "Engineering Dark"],
        index=0,
        key="theme_mode",
    )

    compact_mode = st.toggle(
        "Compact Result Cards",
        value=False,
        key="compact_mode",
    )

    show_static_plots = st.toggle(
        "Show Static Engineering Plots",
        value=False,
        key="show_static_plots",
    )

    show_startup_panel = st.toggle(
        "Show Startup Status",
        value=True,
        key="show_startup_panel",
    )

T = LANGUAGES[language_name]




# =============================================================================
# CONSTANTS
# =============================================================================

ANALYSIS_AXLE = "Axle Whine Order Analysis"
ANALYSIS_TRANSFER_CASE = "Transfer Case Gear Mesh Analysis"

CHANNEL_NAMES = ["ChA", "ChB", "ChC"]

MAX_FILE_SIZE_MB = 500
MAX_ROWS = 3_000_000
G_TO_MS2 = 9.80665


# =============================================================================
# ANALYSIS ENGINE IMPORTS
# =============================================================================

try:
    from order_analysis import (
        read_xlsx_numeric as axle_read_xlsx_numeric,
        angular_resample as axle_angular_resample,
        order_map as axle_order_map,
        extract_order_vs_rpm as axle_extract_order_vs_rpm,
    )
except Exception:
    st.error("order_analysis.py could not be loaded.")
    st.code(traceback.format_exc())
    st.stop()

try:
    from transfer_case_analysis import (
        read_xlsx_numeric as tc_read_xlsx_numeric,
        angular_resample as tc_angular_resample,
        order_map as tc_order_map,
        analyze_transfer_case_orders,
        TRANSFER_CASE_ORDERS,
    )
except Exception:
    st.error("transfer_case_analysis.py could not be loaded.")
    st.code(traceback.format_exc())
    st.stop()




# =============================================================================
# SIDEBAR NAVIGATION
# =============================================================================


# =============================================================================
# GLOBAL COMMAND BAR
# =============================================================================
top_bar = st.container()
with top_bar:
    c1, c2, c3, c4 = st.columns([2.5,1,1,1])
    with c1:
        st.markdown("## 🔧 NVH Engineering Suite")
    with c2:
        st.button(
            "🏠 Home",
            width="stretch",
            disabled=True,
            key="command_home",
        )
    with c3:
        st.button(
            "📄 Report",
            width="stretch",
            disabled=not st.session_state.get(
                "analysis_completed",
                False,
            ),
            key="command_report",
        )
    with c4:
        st.button(
            "⬇ Export",
            width="stretch",
            disabled="excel_report" not in st.session_state,
            key="command_export",
        )

if show_startup_panel:
    st.markdown(
        f"""
<div class="startup-panel">
    <div class="startup-title">✓ {T['startup']}</div>
    <div class="startup-text">
        Axle Whine Engine: Ready · Transfer Case Engine: Ready ·
        Interactive Plotting: Ready · Excel Reporting: Ready ·
        PDF Reporting: {'Ready' if REPORTLAB_AVAILABLE else 'Optional package required'}
    </div>
</div>
""",
        unsafe_allow_html=True,
    )

with st.sidebar:
    st.markdown(
        """
<div class="sidebar-brand">
    <div class="sidebar-brand-title">NVH Engineering Suite</div>
    <div class="sidebar-brand-subtitle">
        Order tracking, target evaluation and engineering reporting workspace
    </div>
</div>
""",
        unsafe_allow_html=True,
    )

    st.markdown(
        '<div class="sidebar-section-label">Workspace Navigation</div>',
        unsafe_allow_html=True,
    )

    st.markdown(
        """
<a class="sidebar-nav-link" href="#vehicle-section">
    01 · Vehicle Information
</a>
<a class="sidebar-nav-link" href="#measurement-section">
    02 · Measurement Data
</a>
<a class="sidebar-nav-link" href="#configuration-section">
    03 · Analysis Configuration
</a>
<a class="sidebar-nav-link" href="#readiness-section">
    04 · Analysis Readiness
</a>
<a class="sidebar-nav-link" href="#results-section">
    05 · Results Dashboard
</a>
<a class="sidebar-nav-link" href="#system-section">
    06 · System Status
</a>
""",
        unsafe_allow_html=True,
    )

    st.markdown(
        '<div class="sidebar-section-label">Module Status</div>',
        unsafe_allow_html=True,
    )

    st.markdown(
        """
<div class="sidebar-status-card">
    <div class="sidebar-status-title">Axle Whine Engine</div>
    <div class="sidebar-status-value">Ready · 10th / 20th order</div>
</div>
<div class="sidebar-status-card">
    <div class="sidebar-status-title">Transfer Case Engine</div>
    <div class="sidebar-status-value">Ready · 63 / 85.05 orders</div>
</div>
<div class="sidebar-status-card">
    <div class="sidebar-status-title">Reporting</div>
    <div class="sidebar-status-value">Excel export enabled</div>
</div>
""",
        unsafe_allow_html=True,
    )

    st.markdown(
        '<div class="sidebar-section-label">Quick Guidance</div>',
        unsafe_allow_html=True,
    )

    st.caption(
        "Complete the workflow from top to bottom. "
        "The Run Analysis button becomes active after all mandatory "
        "inputs are available."
    )


# =============================================================================
# CORPORATE UI THEME
# =============================================================================

st.markdown(
    """
<style>
.stApp {
    background: #f3f6f9;
}

.block-container {
    max-width: 1480px;
    padding-top: 1.3rem;
    padding-bottom: 3rem;
    padding-left: 2rem;
    padding-right: 2rem;
}

html, body, [class*="css"] {
    font-family: Inter, "Segoe UI", Arial, sans-serif;
}

#MainMenu, footer {
    visibility: hidden;
}

header[data-testid="stHeader"] {
    background: transparent;
}

.corporate-header {
    background: linear-gradient(135deg, #0b1f33 0%, #123a63 55%, #1768a6 100%);
    border-radius: 18px;
    padding: 30px 36px;
    margin-bottom: 18px;
    box-shadow: 0 12px 30px rgba(11, 31, 51, 0.18);
    color: #ffffff;
}

.corporate-title {
    font-size: 2.15rem;
    font-weight: 750;
    letter-spacing: -0.02em;
}

.corporate-subtitle {
    color: rgba(255,255,255,0.82);
    font-size: 1rem;
    margin-top: 8px;
}

.corporate-badge {
    display: inline-block;
    margin-top: 14px;
    padding: 6px 12px;
    border-radius: 999px;
    border: 1px solid rgba(255,255,255,0.22);
    background: rgba(255,255,255,0.12);
    font-size: 0.82rem;
    font-weight: 650;
}

.section-title {
    border-left: 4px solid #1768a6;
    padding-left: 12px;
    margin-top: 0.3rem;
    margin-bottom: 0.7rem;
    color: #18324a;
    font-size: 1.08rem;
    font-weight: 750;
}

.section-subtitle {
    color: #6a7d8c;
    font-size: 0.86rem;
    margin-top: 3px;
    font-weight: 400;
}

div[data-testid="stVerticalBlockBorderWrapper"] {
    background: #ffffff;
    border: 1px solid #dce4ea !important;
    border-radius: 15px !important;
    padding: 8px 14px 16px 14px;
    margin-bottom: 16px;
    box-shadow: 0 5px 16px rgba(17, 45, 72, 0.06);
}

div[data-testid="metric-container"] {
    background: #ffffff;
    border: 1px solid #dce4ea;
    border-radius: 12px;
    padding: 14px 16px;
    box-shadow: 0 3px 12px rgba(17, 45, 72, 0.05);
}

div[data-testid="metric-container"] label {
    color: #607585 !important;
    font-weight: 650 !important;
}

div[data-testid="metric-container"] [data-testid="stMetricValue"] {
    color: #17324d;
    font-weight: 750;
}

div[data-baseweb="input"] > div,
div[data-baseweb="select"] > div {
    border-radius: 10px !important;
    border-color: #cbd7df !important;
    background: #ffffff !important;
}

section[data-testid="stFileUploaderDropzone"] {
    background: #ffffff;
    border: 1.5px dashed #9bb2c3;
    border-radius: 14px;
}

.stButton > button {
    min-height: 50px;
    border: none;
    border-radius: 11px;
    background: linear-gradient(135deg, #155d95, #1d78bd);
    color: #ffffff;
    font-weight: 750;
    box-shadow: 0 7px 18px rgba(23, 104, 166, 0.24);
}

.stButton > button:hover {
    color: #ffffff;
    transform: translateY(-1px);
}

.stDownloadButton > button {
    min-height: 46px;
    border-radius: 10px;
    border: 1px solid #1768a6;
    background: #ffffff;
    color: #1768a6;
    font-weight: 750;
}

.stDownloadButton > button:hover {
    background: #1768a6;
    color: #ffffff;
}

button[data-baseweb="tab"] {
    font-weight: 650;
}

button[data-baseweb="tab"][aria-selected="true"] {
    color: #1768a6 !important;
}

div[data-testid="stAlert"],
div[data-testid="stDataFrame"] {
    border-radius: 12px;
}

.workflow-card {
    background: #ffffff;
    border: 1px solid #dce4ea;
    border-radius: 12px;
    padding: 12px 14px;
    margin-bottom: 14px;
    box-shadow: 0 3px 10px rgba(17,45,72,0.04);
}

.workflow-number {
    color: #1768a6;
    font-size: 0.76rem;
    font-weight: 800;
    letter-spacing: 0.08em;
}

.workflow-name {
    color: #18324a;
    font-size: 0.95rem;
    font-weight: 750;
    margin-top: 2px;
}

.info-panel {
    background: #f7fafc;
    border: 1px solid #dce4ea;
    border-radius: 12px;
    padding: 16px;
    min-height: 136px;
}

.info-panel-title {
    color: #18324a;
    font-weight: 750;
    margin-bottom: 8px;
}

.info-panel-body {
    color: #647787;
    font-size: 0.88rem;
    line-height: 1.7;
}
</style>
""",
    unsafe_allow_html=True,
)

st.markdown(
    """
<div class="corporate-header">
    <div class="corporate-title">NVH Analysis Suite</div>
    <div class="corporate-subtitle">
        Axle Whine and Transfer Case Gear Mesh Analysis Platform
    </div>
    <div class="corporate-badge">Engineering Validation Environment</div>
</div>
""",
    unsafe_allow_html=True,
)


if theme_mode == "Engineering Dark":
    st.markdown(
        """
<style>
.stApp {background:#111827;color:#F3F4F6;}
div[data-testid="stVerticalBlockBorderWrapper"]{
background:#1F2937!important;border:1px solid #374151!important;
}
.section-title{color:#F3F4F6!important;border-left-color:#3B82F6!important;}
.section-subtitle{color:#B8C5D1!important;}
div[data-testid="metric-container"]{
background:#1F2937!important;border:1px solid #374151!important;
}
div[data-testid="metric-container"] label,
div[data-testid="metric-container"] [data-testid="stMetricValue"]{
color:#F3F4F6!important;
}
.info-panel{background:#172033!important;border-color:#374151!important;}
.info-panel-title,.info-panel-body{color:#E5E7EB!important;}

.startup-panel{background:linear-gradient(135deg,#FFF 0%,#F4F8FC 100%);
border:1px solid #DCE4EA;border-radius:14px;padding:16px 18px;
margin-bottom:14px;box-shadow:0 5px 14px rgba(17,45,72,.06);}
.startup-title{color:#17324D;font-size:1rem;font-weight:800;}
.startup-text{color:#657887;font-size:.86rem;margin-top:4px;}
.progress-step{border-radius:10px;padding:10px 12px;margin:4px 0;
border:1px solid #DCE4EA;background:#FFF;}
.progress-step-complete{border-left:5px solid #2E8B57;}
.progress-step-active{border-left:5px solid #1768A6;}
.progress-step-pending{border-left:5px solid #AEBCC7;opacity:.76;}
.result-summary-card{border:1px solid #DCE4EA;border-radius:13px;padding:15px;
background:linear-gradient(135deg,#FFF 0%,#F8FAFC 100%);
box-shadow:0 4px 14px rgba(17,45,72,.05);min-height:132px;}
.result-card-title{color:#17324D;font-weight:800;font-size:.92rem;}
.result-card-value{color:#17324D;font-weight:800;font-size:1.35rem;margin-top:8px;}
.result-card-caption{color:#647787;font-size:.78rem;margin-top:5px;}

</style>
""",
        unsafe_allow_html=True,
    )


# =============================================================================
# TARGET DEFINITIONS
# =============================================================================

AXLE_TARGETS = {
    "Diesel": {
        "Front Axle": {
            "rpm": np.array([1000, 1500, 2000, 2500, 3000, 3500, 4000, 4500], dtype=float),
            "amp": np.array([2.5, 2.5, 2.5, 7.5, 7.5, 7.5, 7.5, 7.5], dtype=float),
        },
        "Rear Axle": {
            "rpm": np.array([1000, 1500, 2000, 2500, 3000, 3500, 4000, 4500], dtype=float),
            "amp": np.array([2.5, 2.5, 2.5, 7.5, 7.5, 7.5, 7.5, 7.5], dtype=float),
        },
    },
    "Gasoline": {
        "Front Axle": {
            "rpm": np.array([1000, 1500, 2000, 2500, 3000, 3500, 4000, 4500], dtype=float),
            "amp": np.array([2.5, 2.5, 2.5, 6.25, 10.0, 10.0, 10.0, 10.0], dtype=float),
        },
        "Rear Axle": {
            "rpm": np.array([1000, 1500, 2000, 2500, 3000, 3500, 4000, 4500], dtype=float),
            "amp": np.array([5.0, 5.0, 5.0, 10.0, 12.5, 12.5, 12.5, 12.5], dtype=float),
        },
    },
}


# =============================================================================
# COMMON HELPERS
# =============================================================================

def section_title(title: str, subtitle: Optional[str] = None) -> None:
    subtitle_html = (
        f'<div class="section-subtitle">{subtitle}</div>'
        if subtitle
        else ""
    )
    st.markdown(
        f'<div class="section-title">{title}{subtitle_html}</div>',
        unsafe_allow_html=True,
    )


def build_axle_order_definitions(
    fuel_type: str,
    axle_type: str,
) -> Dict[float, dict]:
    target = AXLE_TARGETS[fuel_type][axle_type]
    return {
        10.0: {
            "label": "10th Order",
            "harmonic": "Base",
            "target_rpm": target["rpm"],
            "target_amp": target["amp"],
        },
        20.0: {
            "label": "20th Order",
            "harmonic": "2nd",
            "target_rpm": target["rpm"],
            "target_amp": target["amp"],
        },
    }


def clear_result_state() -> None:
    result_keys = [
        "analysis_completed",
        "analysis_type_result",
        "time_result",
        "rpm_result",
        "channels_result",
        "curves_by_order",
        "results_by_order",
        "raw_curves_by_order",
        "order_definitions_result",
        "overall_status",
        "vehicle_configuration_result",
        "selected_channel_result",
        "analysis_settings_result",
        "excel_report",
        "pdf_report",
        "vehicle_information",
        "vin_result",
    ]
    for key in result_keys:
        st.session_state.pop(key, None)


def build_input_signature(
    vin: str,
    analysis_type: str,
    fuel_type: str,
    axle_type: str,
    uploaded_file,
    max_order: float,
    order_width: float,
    selected_channel: str,
) -> tuple:
    file_name = None if uploaded_file is None else uploaded_file.name
    file_size = None if uploaded_file is None else int(uploaded_file.size)
    return (
        vin,
        analysis_type,
        fuel_type,
        axle_type,
        file_name,
        file_size,
        float(max_order),
        float(order_width),
        selected_channel,
    )


def convert_csv_g_to_ms2_if_needed(
    headers: list,
    data: np.ndarray,
) -> Tuple[np.ndarray, list]:
    converted_channels = []
    if len(headers) < 4:
        return data, converted_channels

    for column_index in (1, 2, 3):
        header = str(headers[column_index]).strip().lower()
        is_g_unit = (
            "(g)" in header
            or "[g]" in header
            or header.endswith(" g")
        )
        if is_g_unit:
            data[:, column_index] *= G_TO_MS2
            converted_channels.append(str(headers[column_index]))

    return data, converted_channels


def load_measurement_file(
    uploaded_file,
    analysis_type: str,
) -> Tuple[list, np.ndarray]:
    if uploaded_file is None:
        raise ValueError("No measurement file was uploaded.")

    if uploaded_file.size > MAX_FILE_SIZE_MB * 1024 * 1024:
        raise ValueError(
            f"File exceeds the maximum allowed size of {MAX_FILE_SIZE_MB} MB."
        )

    extension = uploaded_file.name.rsplit(".", 1)[-1].lower()
    converted_channels = []

    if extension == "xlsx":
        temporary_path = None
        try:
            with tempfile.NamedTemporaryFile(
                delete=False,
                suffix=".xlsx",
            ) as temporary_file:
                temporary_file.write(uploaded_file.getvalue())
                temporary_path = temporary_file.name

            if analysis_type == ANALYSIS_TRANSFER_CASE:
                headers, data = tc_read_xlsx_numeric(temporary_path)
            else:
                headers, data = axle_read_xlsx_numeric(temporary_path)
        finally:
            if temporary_path and os.path.exists(temporary_path):
                try:
                    os.remove(temporary_path)
                except OSError:
                    pass

    elif extension == "csv":
        uploaded_file.seek(0)
        dataframe = pd.read_csv(
            uploaded_file,
            sep=None,
            engine="python",
        )
        headers = list(dataframe.columns)
        try:
            data = dataframe.to_numpy(dtype=float)
        except (TypeError, ValueError) as error:
            raise ValueError(
                "CSV contains non-numeric values in the measurement columns."
            ) from error

        data, converted_channels = convert_csv_g_to_ms2_if_needed(
            headers,
            data,
        )
    else:
        raise ValueError(
            "Unsupported file format. Please upload an XLSX or CSV file."
        )

    data = np.asarray(data, dtype=float)

    if data.ndim != 2 or data.shape[1] < 5:
        raise ValueError(
            "The file must contain at least five columns in this order: "
            "Time, ChA, ChB, ChC, RPM."
        )

    if data.shape[0] < 10:
        raise ValueError("The measurement file is too short for order analysis.")

    if data.shape[0] > MAX_ROWS:
        raise ValueError(
            f"The dataset exceeds the maximum row limit of {MAX_ROWS:,}."
        )

    if not np.all(np.isfinite(data[:, :5])):
        raise ValueError(
            "The first five columns contain NaN, infinite or non-numeric values."
        )

    if np.any(np.diff(data[:, 0]) < 0):
        raise ValueError("Time values must not decrease.")

    if np.any(data[:, 4] <= 0):
        raise ValueError("RPM values must be positive.")

    if converted_channels:
        st.info(
            "CSV channels converted from g to m/s²: "
            + ", ".join(converted_channels)
        )

    return headers, data


def integrate_positive_area(
    rpm: np.ndarray,
    difference: np.ndarray,
) -> float:
    rpm = np.asarray(rpm, dtype=float)
    difference = np.asarray(difference, dtype=float)
    positive = np.maximum(difference, 0.0)

    if hasattr(np, "trapezoid"):
        return float(np.trapezoid(positive, rpm))

    return float(np.trapz(positive, rpm))


def evaluate_curve_against_target(
    rpm: np.ndarray,
    amplitude: np.ndarray,
    target_rpm: Optional[np.ndarray],
    target_amp: Optional[np.ndarray],
) -> dict:
    rpm = np.asarray(rpm, dtype=float)
    amplitude = np.asarray(amplitude, dtype=float)

    if len(rpm) == 0 or len(amplitude) == 0:
        raise ValueError("No valid order curve was generated.")

    peak_index = int(np.argmax(amplitude))
    peak_rpm = float(rpm[peak_index])
    peak_amplitude = float(amplitude[peak_index])

    if target_rpm is None or target_amp is None:
        return {
            "Peak RPM": peak_rpm,
            "Peak Amplitude [m/s²]": peak_amplitude,
            "Target at Peak RPM [m/s²]": np.nan,
            "Max Margin [m/s²]": np.nan,
            "Max Margin [%]": np.nan,
            "Exceedance Area [m/s²·RPM]": np.nan,
            "Status": "INFO",
        }

    target_curve = np.interp(rpm, target_rpm, target_amp)
    margin_curve = amplitude - target_curve
    margin_index = int(np.argmax(margin_curve))
    target_at_margin = float(target_curve[margin_index])
    max_margin = float(margin_curve[margin_index])

    return {
        "Peak RPM": peak_rpm,
        "Peak Amplitude [m/s²]": peak_amplitude,
        "Target at Peak RPM [m/s²]": float(
            np.interp(peak_rpm, target_rpm, target_amp)
        ),
        "Max Margin [m/s²]": max_margin,
        "Max Margin [%]": (
            max_margin / target_at_margin * 100.0
            if target_at_margin > 0
            else np.nan
        ),
        "Exceedance Area [m/s²·RPM]": integrate_positive_area(
            rpm,
            margin_curve,
        ),
        "Status": (
            "PASS"
            if integrate_positive_area(rpm, margin_curve) <= 1e-9
            else "FAIL"
        ),
    }


# =============================================================================
# AXLE ANALYSIS WRAPPER
# =============================================================================

def analyze_axle_orders(
    time: np.ndarray,
    rpm: np.ndarray,
    channels: Mapping[str, np.ndarray],
    order_definitions: Mapping[float, dict],
    samples_per_rev: int,
    revs_per_block: int,
    overlap: float,
    max_order: float,
    order_width: float,
    rpm_step: float,
    calibration_factor: float,
) -> Tuple[
    Dict[float, dict],
    Dict[float, pd.DataFrame],
    Dict[float, pd.DataFrame],
]:
    highest_order = max(float(order_value) for order_value in order_definitions)
    if max_order < highest_order:
        raise ValueError(
            f"Max Order must be at least {highest_order:.2f}."
        )

    curves_by_order: Dict[float, dict] = {}
    results_by_order: Dict[float, pd.DataFrame] = {}
    raw_curves_by_order: Dict[float, pd.DataFrame] = {}

    for order_value, definition in order_definitions.items():
        order_value = float(order_value)
        channel_curves = {}
        result_rows = []

        for channel_name, signal in channels.items():
            theta_u, signal_u, rpm_u = axle_angular_resample(
                time,
                rpm,
                np.asarray(signal, dtype=float),
                samples_per_rev=samples_per_rev,
            )

            orders, block_rpms, spectrum = axle_order_map(
                theta_u,
                signal_u,
                rpm_u,
                samples_per_rev=samples_per_rev,
                revs_per_block=revs_per_block,
                overlap=overlap,
                max_order=max_order,
            )

            rpm_curve, amplitude_curve = axle_extract_order_vs_rpm(
                orders,
                block_rpms,
                spectrum,
                target_order=order_value,
                width=order_width,
                rpm_step=rpm_step,
                smooth=True,
            )

            rpm_curve = np.asarray(rpm_curve, dtype=float)
            amplitude_curve = (
                np.asarray(amplitude_curve, dtype=float)
                * calibration_factor
            )

            channel_curves[channel_name] = {
                "rpm": rpm_curve,
                "amp": amplitude_curve,
            }

            evaluation = evaluate_curve_against_target(
                rpm_curve,
                amplitude_curve,
                definition["target_rpm"],
                definition["target_amp"],
            )

            result_rows.append(
                {
                    "Order": order_value,
                    "Order Label": definition["label"],
                    "Harmonic": definition["harmonic"],
                    "Channel": channel_name,
                    **evaluation,
                }
            )

        result_dataframe = pd.DataFrame(result_rows)
        first_curve = next(iter(channel_curves.values()))
        common_rpm = np.asarray(first_curve["rpm"], dtype=float)
        curve_dataframe = pd.DataFrame({"RPM": common_rpm})

        for channel_name, curve in channel_curves.items():
            curve_dataframe[channel_name] = np.interp(
                common_rpm,
                curve["rpm"],
                curve["amp"],
            )

        curve_dataframe["Target"] = np.interp(
            common_rpm,
            definition["target_rpm"],
            definition["target_amp"],
        )

        curves_by_order[order_value] = channel_curves
        results_by_order[order_value] = result_dataframe
        raw_curves_by_order[order_value] = curve_dataframe

    return curves_by_order, results_by_order, raw_curves_by_order


# =============================================================================
# PLOTTING
# =============================================================================

def plot_order_comparison(
    order_label: str,
    channel_curves: Mapping[str, dict],
    target_rpm: Optional[np.ndarray],
    target_amp: Optional[np.ndarray],
    vin: str,
    analysis_type: str,
    vehicle_configuration: str,
):
    """
    Create a corporate-style order-vs-RPM comparison plot.
    """
    channel_colors = {
        "ChA": "#1768A6",
        "ChB": "#E67E22",
        "ChC": "#2E8B57",
    }

    figure, axis = plt.subplots(
        figsize=(12.5, 7.2)
    )

    figure.patch.set_facecolor(
        "#F5F7FA"
    )

    axis.set_facecolor(
        "#FFFFFF"
    )

    peak_candidates = []

    for channel_name, curve in channel_curves.items():
        rpm_values = np.asarray(
            curve["rpm"],
            dtype=float,
        )

        amplitude_values = np.asarray(
            curve["amp"],
            dtype=float,
        )

        color = channel_colors.get(
            channel_name,
            "#5B6770",
        )

        axis.plot(
            rpm_values,
            amplitude_values,
            label=channel_name,
            linewidth=2.4,
            color=color,
            solid_capstyle="round",
            zorder=3,
        )

        if len(amplitude_values) > 0:
            peak_index = int(
                np.argmax(
                    amplitude_values
                )
            )

            peak_candidates.append(
                {
                    "channel": channel_name,
                    "rpm": float(
                        rpm_values[
                            peak_index
                        ]
                    ),
                    "amp": float(
                        amplitude_values[
                            peak_index
                        ]
                    ),
                    "color": color,
                }
            )

    if (
        target_rpm is not None
        and target_amp is not None
    ):
        target_rpm_values = np.asarray(
            target_rpm,
            dtype=float,
        )

        target_amp_values = np.asarray(
            target_amp,
            dtype=float,
        )

        axis.plot(
            target_rpm_values,
            target_amp_values,
            label="Target",
            linewidth=3.5,
            color="#C0392B",
            linestyle="--",
            zorder=4,
        )

    if peak_candidates:
        global_peak = max(
            peak_candidates,
            key=lambda item: item["amp"],
        )

        axis.scatter(
            global_peak["rpm"],
            global_peak["amp"],
            s=70,
            color=global_peak["color"],
            edgecolor="#FFFFFF",
            linewidth=1.5,
            zorder=6,
        )

        axis.annotate(
            (
                f"{global_peak['channel']} Peak\n"
                f"{global_peak['amp']:.2f} m/s² @ "
                f"{global_peak['rpm']:.0f} rpm"
            ),
            xy=(
                global_peak["rpm"],
                global_peak["amp"],
            ),
            xytext=(
                18,
                18,
            ),
            textcoords="offset points",
            fontsize=9,
            color="#17324D",
            bbox={
                "boxstyle": "round,pad=0.35",
                "facecolor": "#FFFFFF",
                "edgecolor": "#CBD7DF",
                "alpha": 0.95,
            },
            arrowprops={
                "arrowstyle": "->",
                "color": "#607585",
                "lw": 1.0,
            },
        )

    axis.set_xlabel(
        "Engine Speed [rpm]",
        fontsize=11,
        fontweight="bold",
        color="#30485C",
        labelpad=10,
    )

    axis.set_ylabel(
        "Order Amplitude [m/s²]",
        fontsize=11,
        fontweight="bold",
        color="#30485C",
        labelpad=10,
    )

    axis.set_title(
        f"{order_label} — Order Response",
        loc="left",
        fontsize=15,
        fontweight="bold",
        color="#17324D",
        pad=24,
    )

    axis.text(
        0.0,
        1.015,
        (
            f"VIN: {vin}  |  "
            f"{analysis_type}  |  "
            f"{vehicle_configuration}"
        ),
        transform=axis.transAxes,
        fontsize=9.5,
        color="#6A7D8C",
        va="bottom",
        ha="left",
    )

    axis.grid(
        True,
        which="major",
        linestyle="-",
        linewidth=0.7,
        color="#DCE4EA",
        alpha=0.85,
        zorder=0,
    )

    axis.minorticks_on()

    axis.grid(
        True,
        which="minor",
        linestyle=":",
        linewidth=0.45,
        color="#E9EEF2",
        alpha=0.7,
        zorder=0,
    )

    axis.tick_params(
        axis="both",
        labelsize=9.5,
        colors="#536979",
    )

    axis.spines["top"].set_visible(
        False
    )

    axis.spines["right"].set_visible(
        False
    )

    axis.spines["left"].set_color(
        "#AEBCC7"
    )

    axis.spines["bottom"].set_color(
        "#AEBCC7"
    )

    axis.legend(
        loc="upper left",
        frameon=True,
        fancybox=True,
        framealpha=0.95,
        facecolor="#FFFFFF",
        edgecolor="#DCE4EA",
        fontsize=9.5,
        ncol=4,
    )

    axis.margins(
        x=0.02,
        y=0.08,
    )

    figure.tight_layout(
        pad=2.2
    )

    return figure


def create_order_map_figure(
    time: np.ndarray,
    rpm: np.ndarray,
    signal: np.ndarray,
    selected_channel: str,
    analysis_type: str,
    vin: str,
    samples_per_rev: int,
    revs_per_block: int,
    overlap: float,
    max_order: float,
    calibration_factor: float,
):
    """
    Create a corporate-style order map / waterfall plot.
    """
    engine_angular_resample = (
        tc_angular_resample
        if analysis_type == ANALYSIS_TRANSFER_CASE
        else axle_angular_resample
    )

    engine_order_map = (
        tc_order_map
        if analysis_type == ANALYSIS_TRANSFER_CASE
        else axle_order_map
    )

    theta_u, signal_u, rpm_u = engine_angular_resample(
        time,
        rpm,
        signal,
        samples_per_rev=samples_per_rev,
    )

    orders, block_rpms, spectrum = engine_order_map(
        theta_u,
        signal_u,
        rpm_u,
        samples_per_rev=samples_per_rev,
        revs_per_block=revs_per_block,
        overlap=overlap,
        max_order=max_order,
    )

    orders = np.asarray(
        orders,
        dtype=float,
    )

    block_rpms = np.asarray(
        block_rpms,
        dtype=float,
    )

    spectrum = np.asarray(
        spectrum,
        dtype=float,
    )

    if spectrum.ndim != 2:
        raise ValueError(
            "Order spectrum must be two-dimensional."
        )

    if spectrum.shape[0] != len(block_rpms):
        raise ValueError(
            "Order spectrum row count does not match the RPM vector."
        )

    if spectrum.shape[1] != len(orders):
        raise ValueError(
            "Order spectrum column count does not match the order axis."
        )

    sort_index = np.argsort(
        block_rpms,
        kind="stable",
    )

    sorted_rpm = block_rpms[
        sort_index
    ]

    sorted_spectrum = spectrum[
        sort_index,
        :
    ]

    decibels = 20.0 * np.log10(
        np.maximum(
            sorted_spectrum
            * calibration_factor,
            1e-12,
        )
    )

    figure, axis = plt.subplots(
        figsize=(12.5, 7.2)
    )

    figure.patch.set_facecolor(
        "#F5F7FA"
    )

    axis.set_facecolor(
        "#FFFFFF"
    )

    image = axis.imshow(
        decibels,
        aspect="auto",
        origin="lower",
        extent=[
            float(orders[0]),
            float(orders[-1]),
            float(sorted_rpm[0]),
            float(sorted_rpm[-1]),
        ],
        interpolation="nearest",
        cmap="turbo",
    )

    colorbar = figure.colorbar(
        image,
        ax=axis,
        pad=0.02,
        fraction=0.035,
    )

    colorbar.set_label(
        "Amplitude [dB re 1 m/s²]",
        fontsize=10,
        fontweight="bold",
        color="#30485C",
        labelpad=10,
    )

    colorbar.ax.tick_params(
        labelsize=9,
        colors="#536979",
    )

    axis.set_xlabel(
        "Order",
        fontsize=11,
        fontweight="bold",
        color="#30485C",
        labelpad=10,
    )

    axis.set_ylabel(
        "Engine Speed [rpm]",
        fontsize=11,
        fontweight="bold",
        color="#30485C",
        labelpad=10,
    )

    axis.set_title(
        f"Order Map / Waterfall — {selected_channel}",
        loc="left",
        fontsize=15,
        fontweight="bold",
        color="#17324D",
        pad=24,
    )

    axis.text(
        0.0,
        1.015,
        (
            f"VIN: {vin}  |  "
            f"{analysis_type}  |  "
            f"Max Order: {max_order:.0f}"
        ),
        transform=axis.transAxes,
        fontsize=9.5,
        color="#6A7D8C",
        va="bottom",
        ha="left",
    )

    axis.tick_params(
        axis="both",
        labelsize=9.5,
        colors="#536979",
    )

    axis.spines["top"].set_visible(
        False
    )

    axis.spines["right"].set_visible(
        False
    )

    axis.spines["left"].set_color(
        "#AEBCC7"
    )

    axis.spines["bottom"].set_color(
        "#AEBCC7"
    )

    figure.tight_layout(
        pad=2.0
    )

    return figure


# =============================================================================
# EXCEL REPORT
# =============================================================================

def format_comparison_sheet(writer, sheet_name: str) -> None:
    worksheet = writer.book[sheet_name]

    fills = {
        "PASS": PatternFill(
            start_color="C6EFCE",
            end_color="C6EFCE",
            fill_type="solid",
        ),
        "FAIL": PatternFill(
            start_color="FFC7CE",
            end_color="FFC7CE",
            fill_type="solid",
        ),
        "INFO": PatternFill(
            start_color="D9EAF7",
            end_color="D9EAF7",
            fill_type="solid",
        ),
    }

    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid",
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    status_column = next(
        (
            cell.column
            for cell in worksheet[1]
            if cell.value == "Status"
        ),
        None,
    )

    if status_column is not None:
        for row_index in range(2, worksheet.max_row + 1):
            status = worksheet.cell(
                row=row_index,
                column=status_column,
            ).value
            row_fill = fills.get(status)
            if row_fill is not None:
                for column_index in range(1, worksheet.max_column + 1):
                    worksheet.cell(
                        row=row_index,
                        column=column_index,
                    ).fill = row_fill

    for column_cells in worksheet.columns:
        worksheet.column_dimensions[
            column_cells[0].column_letter
        ].width = 22


def format_curve_sheet(writer, sheet_name: str) -> None:
    worksheet = writer.book[sheet_name]
    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid",
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for column_cells in worksheet.columns:
        worksheet.column_dimensions[
            column_cells[0].column_letter
        ].width = 16


def create_curve_plot_png(
    curve_dataframe: pd.DataFrame,
    order_label: str,
    vin: str,
    analysis_type: str,
    vehicle_configuration: str,
) -> BytesIO:
    figure, axis = plt.subplots(figsize=(14, 8))

    for channel_name in CHANNEL_NAMES:
        if channel_name in curve_dataframe.columns:
            axis.plot(
                curve_dataframe["RPM"],
                curve_dataframe[channel_name],
                label=channel_name,
                linewidth=2,
            )

    if "Target" in curve_dataframe.columns:
        axis.plot(
            curve_dataframe["RPM"],
            curve_dataframe["Target"],
            label="Target Curve",
            linewidth=5,
        )

    axis.set_title(
        f"{order_label} vs RPM | VIN: {vin} | "
        f"{analysis_type} | {vehicle_configuration}",
        fontsize=16,
    )
    axis.set_xlabel("RPM", fontsize=13)
    axis.set_ylabel("Order Amplitude [m/s²]", fontsize=13)
    axis.grid(True, alpha=0.3)
    axis.legend(loc="upper right", fontsize=12)
    figure.tight_layout()

    image_buffer = BytesIO()
    figure.savefig(
        image_buffer,
        format="png",
        dpi=180,
        bbox_inches="tight",
    )
    plt.close(figure)
    image_buffer.seek(0)
    return image_buffer


def make_excel_report(
    vehicle_information: dict,
    results_by_order: Mapping[float, pd.DataFrame],
    curves_by_order: Mapping[float, pd.DataFrame],
    order_definitions: Mapping[float, dict],
) -> BytesIO:
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame([vehicle_information]).to_excel(
            writer,
            sheet_name="Vehicle Info",
            index=False,
        )

        for order_value, result_dataframe in results_by_order.items():
            comparison_sheet = (
                f"{str(order_value).replace('.', '_')} Comparison"
            )[:31]
            result_dataframe.to_excel(
                writer,
                sheet_name=comparison_sheet,
                index=False,
            )
            format_comparison_sheet(writer, comparison_sheet)

        for order_value, curve_dataframe in curves_by_order.items():
            curve_sheet = (
                f"{str(order_value).replace('.', '_')} Curves"
            )[:31]
            curve_dataframe.to_excel(
                writer,
                sheet_name=curve_sheet,
                index=False,
            )
            format_curve_sheet(writer, curve_sheet)

            image_buffer = create_curve_plot_png(
                curve_dataframe=curve_dataframe,
                order_label=order_definitions[order_value]["label"],
                vin=vehicle_information["VIN"],
                analysis_type=vehicle_information["Analysis Type"],
                vehicle_configuration=vehicle_information[
                    "Vehicle Configuration"
                ],
            )

            image = XLImage(image_buffer)
            image.width = 900
            image.height = 520
            writer.book[curve_sheet].add_image(image, "G2")

    output.seek(0)
    return output


# =============================================================================
# PDF REPORT GENERATION
# =============================================================================

def make_pdf_report(
    vehicle_information: dict,
    results_by_order: Mapping[float, pd.DataFrame],
) -> Optional[BytesIO]:
    if not REPORTLAB_AVAILABLE:
        return None

    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="NVH Engineering Analysis Report",
        author="NVH Engineering Suite",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CorporateTitle",
        parent=styles["Title"],
        alignment=TA_CENTER,
        textColor=colors.HexColor("#17324D"),
        fontSize=20,
        leading=24,
        spaceAfter=10,
    )
    heading_style = ParagraphStyle(
        "CorporateHeading",
        parent=styles["Heading2"],
        textColor=colors.HexColor("#1768A6"),
        fontSize=13,
        leading=16,
        spaceBefore=8,
        spaceAfter=6,
    )

    story = [
        Paragraph("NVH Engineering Analysis Report", title_style),
        Paragraph(
            f"VIN: {vehicle_information.get('VIN', 'N/A')} | "
            f"{vehicle_information.get('Analysis Type', 'N/A')} | "
            f"{vehicle_information.get('Vehicle Configuration', 'N/A')}",
            styles["Normal"],
        ),
        Spacer(1, 8),
        Paragraph("Executive Summary", heading_style),
    ]

    summary_data = [["Field", "Value"]]
    for key in [
        "VIN","Analysis Type","Fuel Type","Vehicle Configuration",
        "Target Orders","Order Width","RPM Step","Samples per Rev",
        "Revs per Block","Overlap","Max Order","Overall Assessment",
    ]:
        summary_data.append([key, str(vehicle_information.get(key, "N/A"))])

    summary_table = Table(summary_data, colWidths=[55 * mm, 190 * mm], repeatRows=1)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#17324D")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#CBD7DF")),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F3F6F9")]),
        ("FONTSIZE",(0,0),(-1,-1),8.5),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.extend([summary_table, PageBreak()])

    for order_value, dataframe in results_by_order.items():
        story.append(Paragraph(f"{order_value:.2f} Order Results", heading_style))
        cols = [c for c in [
            "Order Label","Harmonic","Channel","Peak RPM",
            "Peak Amplitude [m/s²]","Target at Peak RPM [m/s²]",
            "Max Margin [m/s²]","Max Margin [%]",
            "Exceedance Area [m/s²·RPM]","Status",
        ] if c in dataframe.columns]
        data = [cols]
        for _, row in dataframe[cols].iterrows():
            data.append([
                f"{v:.3f}" if isinstance(v,(float,np.floating)) and np.isfinite(v) else str(v)
                for v in row.tolist()
            ])
        table = Table(data, repeatRows=1, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1768A6")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#CBD7DF")),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F3F6F9")]),
            ("FONTSIZE",(0,0),(-1,-1),7.2),
        ]))
        story.extend([table, Spacer(1, 10)])

    document.build(story)
    output.seek(0)
    return output



# =============================================================================
# UX WORKFLOW WIZARD
# =============================================================================

def render_progress_wizard(
    vin_valid: bool,
    uploaded_file,
    configuration_selected: bool,
    analysis_completed: bool,
) -> None:
    steps = [
        ("Vehicle Information", vin_valid, "Validate the 17-character VIN."),
        ("Measurement Import", uploaded_file is not None, "Upload measurement data."),
        ("Signal Configuration", configuration_selected, "Confirm module settings."),
        ("Order Tracking", analysis_completed, "Run order extraction."),
        ("Target Evaluation", analysis_completed, "Evaluate target compliance."),
        (
            "Report Generation",
            analysis_completed and "excel_report" in st.session_state,
            "Generate Excel and PDF reports.",
        ),
    ]

    with st.expander("Analysis Workflow", expanded=not analysis_completed):
        for index, (label, completed, description) in enumerate(steps, start=1):
            previous_complete = all(item[1] for item in steps[: index - 1])
            if completed:
                css_class = "progress-step progress-step-complete"
                status_text = "✓ Complete"
            elif previous_complete:
                css_class = "progress-step progress-step-active"
                status_text = "● Current"
            else:
                css_class = "progress-step progress-step-pending"
                status_text = "○ Pending"

            st.markdown(
                f"""
<div class="{css_class}">
<strong>{index:02d} · {label}</strong>
<span style="float:right">{status_text}</span>
<div style="font-size:.80rem;color:#647787;margin-top:4px">{description}</div>
</div>
""",
                unsafe_allow_html=True,
            )



st.markdown('<div id="vehicle-section"></div>', unsafe_allow_html=True)

with st.container(border=True):
    section_title(
        "Vehicle Information",
        "Identify the vehicle and select the required NVH analysis module.",
    )

    vehicle_column, analysis_column, option_1, option_2 = st.columns(
        [1.15, 1.35, 1.0, 1.0]
    )

    with vehicle_column:
        vin_number = st.text_input(
            "VIN Number",
            placeholder="Enter 17-character VIN",
            max_chars=17,
            key="vin_number",
        ).upper().strip()

    vin_valid = bool(re.fullmatch(r"[A-Z0-9]{17}", vin_number))

    with analysis_column:
        analysis_type = st.selectbox(
            "Analysis Type",
            [ANALYSIS_AXLE, ANALYSIS_TRANSFER_CASE],
            disabled=not vin_valid,
            key="analysis_type",
        )

    if analysis_type == ANALYSIS_AXLE:
        with option_1:
            fuel_type = st.selectbox(
                "Fuel Type",
                ["Select fuel type", "Diesel", "Gasoline"],
                disabled=not vin_valid,
                key="fuel_type",
            )

        with option_2:
            axle_type = st.selectbox(
                "Axle Type",
                ["Select axle type", "Front Axle", "Rear Axle"],
                disabled=not vin_valid,
                key="axle_type",
            )

        vehicle_configuration = f"{fuel_type} | {axle_type}"

    else:
        fuel_type = "N/A"
        axle_type = "Transfer Case / 6th Gear"
        vehicle_configuration = "Transfer Case | 6th Gear"

        with option_1:
            st.text_input(
                "Gear",
                value="6th Gear",
                disabled=True,
                key="tc_gear",
            )

        with option_2:
            st.text_input(
                "Component",
                value="Transfer Case",
                disabled=True,
                key="tc_component",
            )

    if vin_number and not vin_valid:
        st.error(
            "VIN must contain exactly 17 letters or numbers."
        )
    elif vin_valid:
        st.success("Vehicle identification completed.")


st.markdown('<div id="measurement-section"></div>', unsafe_allow_html=True)

with st.container(border=True):
    section_title(
        "Measurement Data",
        "Upload Time, ChA, ChB, ChC and RPM measurement data.",
    )

    upload_column, structure_column = st.columns([1.7, 1.0])

    with upload_column:
        uploaded_file = st.file_uploader(
            "Upload Measurement File",
            type=["xlsx", "csv"],
            disabled=not vin_valid,
            key="measurement_file",
            help=(
                "Expected first five columns: "
                "Time, ChA, ChB, ChC, RPM."
            ),
        )

    with structure_column:
        st.markdown(
            """
<div class="info-panel">
    <div class="info-panel-title">Required Data Structure</div>
    <div class="info-panel-body">
        • Time [s]<br>
        • ChA vibration<br>
        • ChB vibration<br>
        • ChC vibration<br>
        • RPM
    </div>
</div>
""",
            unsafe_allow_html=True,
        )

    if uploaded_file is not None:
        with upload_column:
            st.success(
                "✓ Measurement file loaded successfully."
            )


if analysis_type == ANALYSIS_AXLE:
    configuration_selected = (
        fuel_type != "Select fuel type"
        and axle_type != "Select axle type"
    )

    if configuration_selected:
        order_definitions = build_axle_order_definitions(
            fuel_type,
            axle_type,
        )
    else:
        order_definitions = {}

    fixed_samples_per_rev = 512
    fixed_revs_per_block = 8
    fixed_overlap = 0.75
    fixed_rpm_step = 10.0
    fixed_calibration_factor = 1.0
    minimum_max_order = 20
    default_max_order = 30

else:
    configuration_selected = True
    order_definitions = TRANSFER_CASE_ORDERS
    fixed_samples_per_rev = 512
    fixed_revs_per_block = 20
    fixed_overlap = 0.75
    fixed_rpm_step = 10.0
    fixed_calibration_factor = 1.0
    minimum_max_order = 171
    default_max_order = 200


# =============================================================================
# ANALYSIS CONFIGURATION
# (Hidden from user interface)
# =============================================================================

# Standard engineering settings are applied automatically.

can_continue = (
    vin_valid
    and configuration_selected
    and uploaded_file is not None
)


render_progress_wizard(
    vin_valid=vin_valid,
    uploaded_file=uploaded_file,
    configuration_selected=configuration_selected,
    analysis_completed=st.session_state.get("analysis_completed", False),
)

st.markdown('<div id="readiness-section"></div>', unsafe_allow_html=True)

with st.container(border=True):
    section_title(
        "Analysis Readiness",
        "Confirm the current configuration before running the analysis.",
    )

    readiness_columns = st.columns(4)
    readiness_columns[0].metric(
        "VIN",
        vin_number if vin_valid else "Not ready",
    )
    readiness_columns[1].metric(
        "Analysis",
        (
            "Axle Whine"
            if analysis_type == ANALYSIS_AXLE
            else "Transfer Case"
        ),
    )
    readiness_columns[2].metric(
        "Configuration",
        vehicle_configuration,
    )
    readiness_columns[3].metric(
        "Input Status",
        "Ready" if can_continue else "Incomplete",
    )

    if can_continue:
        st.success(
            "All required inputs are available. "
            "The analysis can be started."
        )
    elif not vin_valid:
        st.warning("Enter a valid 17-character VIN.")
    elif analysis_type == ANALYSIS_AXLE and not configuration_selected:
        st.warning("Select the fuel and axle configuration.")
    else:
        st.warning("Upload a measurement file to continue.")


current_signature = build_input_signature(
    vin=vin_number,
    analysis_type=analysis_type,
    fuel_type=fuel_type,
    axle_type=axle_type,
    uploaded_file=uploaded_file,
    max_order=max_order,
    order_width=order_width,
    selected_channel=selected_channel,
)

previous_signature = st.session_state.get("input_signature")
if (
    previous_signature is not None
    and previous_signature != current_signature
):
    clear_result_state()

st.session_state["input_signature"] = current_signature


# =============================================================================
# RUN ANALYSIS
# =============================================================================

if st.button(
    T["run"],
    type="primary",
    width="stretch",
    disabled=not can_continue,
    key="run_analysis",
):
    try:
        progress = st.progress(0, text="Reading measurement data...")

        _, data = load_measurement_file(
            uploaded_file,
            analysis_type,
        )

        time = np.asarray(data[:, 0], dtype=float)
        rpm = np.asarray(data[:, 4], dtype=float)
        channels = {
            "ChA": np.asarray(data[:, 1], dtype=float),
            "ChB": np.asarray(data[:, 2], dtype=float),
            "ChC": np.asarray(data[:, 3], dtype=float),
        }

        progress.progress(25, text="Preparing angular-domain signals...")

        if analysis_type == ANALYSIS_TRANSFER_CASE:
            progress.progress(45, text="Running Transfer Case order analysis...")
            (
                curves_by_order,
                results_by_order,
                raw_curves_by_order,
            ) = analyze_transfer_case_orders(
                time=time,
                rpm=rpm,
                channels=channels,
                order_definitions=order_definitions,
                samples_per_rev=fixed_samples_per_rev,
                revs_per_block=fixed_revs_per_block,
                overlap=fixed_overlap,
                max_order=max_order,
                order_width=order_width,
                rpm_step=fixed_rpm_step,
                calibration_factor=fixed_calibration_factor,
            )
        else:
            progress.progress(45, text="Running Axle Whine order analysis...")
            (
                curves_by_order,
                results_by_order,
                raw_curves_by_order,
            ) = analyze_axle_orders(
                time=time,
                rpm=rpm,
                channels=channels,
                order_definitions=order_definitions,
                samples_per_rev=fixed_samples_per_rev,
                revs_per_block=fixed_revs_per_block,
                overlap=fixed_overlap,
                max_order=max_order,
                order_width=order_width,
                rpm_step=fixed_rpm_step,
                calibration_factor=fixed_calibration_factor,
            )

        progress.progress(75, text="Evaluating target compliance...")

        statuses = []
        for result_dataframe in results_by_order.values():
            evaluated = result_dataframe[
                result_dataframe["Status"] != "INFO"
            ]
            statuses.extend(evaluated["Status"].tolist())

        if not statuses:
            overall_status = "INFO"
        elif any(status == "FAIL" for status in statuses):
            overall_status = "FAIL"
        else:
            overall_status = "PASS"

        vehicle_information = {
            "VIN": vin_number,
            "Analysis Type": analysis_type,
            "Fuel Type": fuel_type,
            "Vehicle Configuration": vehicle_configuration,
            "Target Orders": ", ".join(
                str(order_value)
                for order_value in order_definitions
            ),
            "Order Width": order_width,
            "RPM Step": fixed_rpm_step,
            "Samples per Rev": fixed_samples_per_rev,
            "Revs per Block": fixed_revs_per_block,
            "Overlap": fixed_overlap,
            "Calibration Factor": fixed_calibration_factor,
            "Max Order": max_order,
            "Overall Assessment": overall_status,
        }

        progress.progress(88, text="Generating Excel report...")

        excel_report = make_excel_report(
            vehicle_information,
            results_by_order,
            raw_curves_by_order,
            order_definitions,
        )

        try:
            pdf_report = make_pdf_report(
                vehicle_information=vehicle_information,
                results_by_order=results_by_order,
            )
        except Exception:
            pdf_report = None

        st.session_state.update(
            {
                "analysis_completed": True,
                "analysis_type_result": analysis_type,
                "time_result": time,
                "rpm_result": rpm,
                "channels_result": channels,
                "curves_by_order": curves_by_order,
                "results_by_order": results_by_order,
                "raw_curves_by_order": raw_curves_by_order,
                "order_definitions_result": order_definitions,
                "overall_status": overall_status,
                "vehicle_configuration_result": vehicle_configuration,
                "selected_channel_result": selected_channel,
                "analysis_settings_result": {
                    "samples_per_rev": fixed_samples_per_rev,
                    "revs_per_block": fixed_revs_per_block,
                    "overlap": fixed_overlap,
                    "max_order": max_order,
                    "calibration_factor": fixed_calibration_factor,
                },
                "excel_report": excel_report,
                "pdf_report": pdf_report,
                "vehicle_information": vehicle_information,
                "vin_result": vin_number,
            }
        )

        progress.progress(100, text="Analysis completed.")
        st.success("Analysis completed successfully.")

    except Exception as error:
        st.session_state["analysis_completed"] = False
        st.error("An error occurred while running the analysis.")
        st.exception(error)


# =============================================================================
# RESULTS
# =============================================================================

def determine_order_status(result_dataframe: pd.DataFrame) -> str:
    evaluated = result_dataframe[
        result_dataframe["Status"] != "INFO"
    ]
    if len(evaluated) == 0:
        return "INFO"
    if (evaluated["Status"] == "PASS").all():
        return "PASS"
    return "FAIL"


def render_analysis_dashboard(
    result_tables: Mapping[float, pd.DataFrame],
    overall_status: str,
    vin: str,
    analysis_type: str,
    vehicle_configuration: str,
) -> None:
    """
    Render the executive KPI dashboard for completed analyses.
    """
    combined_frames = []

    for order_value, result_dataframe in result_tables.items():
        if result_dataframe is None or result_dataframe.empty:
            continue

        frame = result_dataframe.copy()

        if "Order" not in frame.columns:
            frame["Order"] = float(order_value)

        combined_frames.append(frame)

    if not combined_frames:
        st.warning(
            "No result data are available for the executive dashboard."
        )
        return

    combined_results = pd.concat(
        combined_frames,
        ignore_index=True,
    )

    peak_column = "Peak Amplitude [m/s²]"

    if (
        peak_column in combined_results.columns
        and combined_results[peak_column].notna().any()
    ):
        peak_index = (
            combined_results[peak_column]
            .astype(float)
            .idxmax()
        )

        peak_row = combined_results.loc[peak_index]

        maximum_peak = float(
            peak_row[peak_column]
        )

        peak_rpm = float(
            peak_row.get(
                "Peak RPM",
                np.nan,
            )
        )

        worst_channel = str(
            peak_row.get(
                "Channel",
                "N/A",
            )
        )

        peak_order = float(
            peak_row.get(
                "Order",
                np.nan,
            )
        )

    else:
        maximum_peak = np.nan
        peak_rpm = np.nan
        worst_channel = "N/A"
        peak_order = np.nan

    margin_column = "Max Margin [m/s²]"

    if (
        margin_column in combined_results.columns
        and combined_results[margin_column].notna().any()
    ):
        margin_rows = combined_results[
            combined_results[margin_column].notna()
        ]

        margin_index = (
            margin_rows[margin_column]
            .astype(float)
            .idxmax()
        )

        margin_row = margin_rows.loc[
            margin_index
        ]

        worst_order = float(
            margin_row.get(
                "Order",
                np.nan,
            )
        )

        maximum_margin = float(
            margin_row[
                margin_column
            ]
        )

    else:
        worst_order = peak_order
        maximum_margin = np.nan

    def format_order_value(
        value: float,
    ) -> str:
        if not np.isfinite(value):
            return "N/A"

        if abs(value - round(value)) < 1e-9:
            return f"{value:.0f}"

        return f"{value:.2f}"

    peak_display = (
        f"{maximum_peak:.2f} m/s²"
        if np.isfinite(maximum_peak)
        else "N/A"
    )

    peak_rpm_display = (
        f"{peak_rpm:.0f} rpm"
        if np.isfinite(peak_rpm)
        else "N/A"
    )

    margin_display = (
        f"{maximum_margin:.2f} m/s²"
        if np.isfinite(maximum_margin)
        else "N/A"
    )

    evaluated_orders = ", ".join(
        format_order_value(
            float(order_value)
        )
        for order_value in sorted(
            result_tables.keys()
        )
    )

    analysis_display = (
        "Axle Whine"
        if analysis_type == ANALYSIS_AXLE
        else "Transfer Case"
    )

    with st.container(border=True):
        section_title(
            "Executive Analysis Dashboard",
            (
                "Key order-analysis indicators and the most "
                "critical measured condition."
            ),
        )

        first_row = st.columns(4)

        first_row[0].metric(
            "Overall Status",
            overall_status,
        )

        first_row[1].metric(
            "Worst Order",
            format_order_value(
                worst_order
            ),
        )

        first_row[2].metric(
            "Worst Channel",
            worst_channel,
        )

        first_row[3].metric(
            "Maximum Peak",
            peak_display,
        )

        second_row = st.columns(4)

        second_row[0].metric(
            "Peak RPM",
            peak_rpm_display,
        )

        second_row[1].metric(
            "Maximum Margin",
            margin_display,
        )

        second_row[2].metric(
            "Evaluated Orders",
            evaluated_orders,
        )

        second_row[3].metric(
            "Analysis Module",
            analysis_display,
        )

        summary_row = st.columns(2)

        summary_row[0].metric(
            "VIN",
            vin,
        )

        summary_row[1].metric(
            "Vehicle Configuration",
            vehicle_configuration,
        )

        if overall_status == "PASS":
            st.success(
                "Overall Assessment: PASS — "
                "all evaluated channels remain within "
                "the defined targets."
            )

        elif overall_status == "FAIL":
            st.error(
                "Overall Assessment: FAIL — "
                f"the most critical condition is "
                f"{format_order_value(worst_order)} order "
                f"on {worst_channel}."
            )

        else:
            st.info(
                "Overall Assessment: INFO — "
                "the displayed harmonics do not contain "
                "a defined target."
            )


def render_interactive_order_chart(
    order_label: str,
    channel_curves: Mapping[str, dict],
    target_rpm: Optional[np.ndarray],
    target_amp: Optional[np.ndarray],
    vin: str,
    analysis_type: str,
    vehicle_configuration: str,
) -> None:
    """
    Render a Plotly-first engineering chart with unified hover,
    zoom, pan, legend control and image export.

    When Plotly is unavailable, the app falls back to Altair so the
    analysis workflow remains usable.
    """
    channel_colors = {
        "ChA": "#1768A6",
        "ChB": "#E67E22",
        "ChC": "#2E8B57",
    }

    valid_curves = {}

    for channel_name, curve in channel_curves.items():
        rpm_values = np.asarray(
            curve["rpm"],
            dtype=float,
        )

        amplitude_values = np.asarray(
            curve["amp"],
            dtype=float,
        )

        valid_mask = (
            np.isfinite(rpm_values)
            & np.isfinite(amplitude_values)
        )

        if np.any(valid_mask):
            valid_curves[channel_name] = {
                "rpm": rpm_values[valid_mask],
                "amp": amplitude_values[valid_mask],
            }

    if not valid_curves:
        st.warning(
            "No valid curve data are available for the interactive chart."
        )
        return

    if go is not None:
        figure = go.Figure()

        for channel_name, curve in valid_curves.items():
            figure.add_trace(
                go.Scatter(
                    x=curve["rpm"],
                    y=curve["amp"],
                    mode="lines",
                    name=channel_name,
                    line={
                        "color": channel_colors.get(
                            channel_name,
                            "#5B6770",
                        ),
                        "width": 2.6,
                    },
                    hovertemplate=(
                        f"<b>{channel_name}</b><br>"
                        "RPM: %{x:.0f}<br>"
                        "Amplitude: %{y:.3f} m/s²"
                        "<extra></extra>"
                    ),
                )
            )

        if (
            target_rpm is not None
            and target_amp is not None
        ):
            target_rpm_values = np.asarray(
                target_rpm,
                dtype=float,
            )

            target_amp_values = np.asarray(
                target_amp,
                dtype=float,
            )

            target_mask = (
                np.isfinite(target_rpm_values)
                & np.isfinite(target_amp_values)
            )

            if np.any(target_mask):
                figure.add_trace(
                    go.Scatter(
                        x=target_rpm_values[target_mask],
                        y=target_amp_values[target_mask],
                        mode="lines",
                        name="Target",
                        line={
                            "color": "#C0392B",
                            "width": 3.5,
                            "dash": "dash",
                        },
                        hovertemplate=(
                            "<b>Target</b><br>"
                            "RPM: %{x:.0f}<br>"
                            "Target: %{y:.3f} m/s²"
                            "<extra></extra>"
                        ),
                    )
                )

        peak_candidates = []

        for channel_name, curve in valid_curves.items():
            peak_index = int(
                np.argmax(
                    curve["amp"]
                )
            )

            peak_candidates.append(
                {
                    "channel": channel_name,
                    "rpm": float(
                        curve["rpm"][
                            peak_index
                        ]
                    ),
                    "amp": float(
                        curve["amp"][
                            peak_index
                        ]
                    ),
                }
            )

        if peak_candidates:
            global_peak = max(
                peak_candidates,
                key=lambda item: item["amp"],
            )

            figure.add_trace(
                go.Scatter(
                    x=[
                        global_peak["rpm"]
                    ],
                    y=[
                        global_peak["amp"]
                    ],
                    mode="markers",
                    name="Global Peak",
                    marker={
                        "size": 11,
                        "color": channel_colors.get(
                            global_peak["channel"],
                            "#17324D",
                        ),
                        "line": {
                            "color": "#FFFFFF",
                            "width": 1.5,
                        },
                    },
                    hovertemplate=(
                        f"<b>{global_peak['channel']} Global Peak</b><br>"
                        "RPM: %{x:.0f}<br>"
                        "Amplitude: %{y:.3f} m/s²"
                        "<extra></extra>"
                    ),
                    showlegend=False,
                )
            )

            figure.add_annotation(
                x=global_peak["rpm"],
                y=global_peak["amp"],
                text=(
                    f"<b>{global_peak['channel']} Peak</b><br>"
                    f"{global_peak['amp']:.2f} m/s² @ "
                    f"{global_peak['rpm']:.0f} rpm"
                ),
                showarrow=True,
                arrowhead=2,
                arrowsize=1,
                arrowwidth=1.2,
                arrowcolor="#607585",
                ax=68,
                ay=-52,
                bgcolor="rgba(255,255,255,0.96)",
                bordercolor="#CBD7DF",
                borderwidth=1,
                borderpad=6,
                font={
                    "color": "#17324D",
                    "size": 11,
                },
            )

        figure.update_layout(
            title={
                "text": (
                    f"<b>{order_label} — Interactive Order Response</b>"
                    f"<br><sup>VIN: {vin} &nbsp; | &nbsp; "
                    f"{analysis_type} &nbsp; | &nbsp; "
                    f"{vehicle_configuration}</sup>"
                ),
                "x": 0.01,
                "xanchor": "left",
                "font": {
                    "size": 18,
                    "color": "#17324D",
                },
            },
            height=590,
            margin={
                "l": 70,
                "r": 30,
                "t": 95,
                "b": 65,
            },
            paper_bgcolor="#F5F7FA",
            plot_bgcolor="#FFFFFF",
            hovermode="x unified",
            hoverlabel={
                "bgcolor": "#FFFFFF",
                "bordercolor": "#CBD7DF",
                "font": {
                    "color": "#17324D",
                    "size": 12,
                },
            },
            legend={
                "orientation": "h",
                "yanchor": "bottom",
                "y": 1.01,
                "xanchor": "right",
                "x": 1.0,
                "font": {
                    "color": "#30485C",
                },
            },
            xaxis={
                "title": {
                    "text": "Engine Speed [rpm]",
                    "font": {
                        "color": "#30485C",
                        "size": 13,
                    },
                },
                "showgrid": True,
                "gridcolor": "#DCE4EA",
                "zeroline": False,
                "linecolor": "#AEBCC7",
                "tickfont": {
                    "color": "#536979",
                },
                "showspikes": True,
                "spikemode": "across",
                "spikesnap": "cursor",
                "spikecolor": "#607585",
                "spikedash": "dot",
                "spikethickness": 1,
                "rangeslider": {
                    "visible": True,
                    "thickness": 0.07,
                },
            },
            yaxis={
                "title": {
                    "text": "Order Amplitude [m/s²]",
                    "font": {
                        "color": "#30485C",
                        "size": 13,
                    },
                },
                "showgrid": True,
                "gridcolor": "#DCE4EA",
                "zeroline": False,
                "linecolor": "#AEBCC7",
                "tickfont": {
                    "color": "#536979",
                },
                "rangemode": "tozero",
            },
            modebar={
                "bgcolor": "rgba(255,255,255,0.75)",
                "color": "#607585",
                "activecolor": "#1768A6",
            },
        )

        figure.update_xaxes(
            showgrid=True,
            gridcolor="#DCE4EA",
        )

        st.plotly_chart(
            figure,
            width="stretch",
            config={
                "displaylogo": False,
                "scrollZoom": True,
                "responsive": True,
                "toImageButtonOptions": {
                    "format": "png",
                    "filename": (
                        f"{vin}_{order_label.replace(' ', '_')}"
                    ),
                    "height": 720,
                    "width": 1280,
                    "scale": 2,
                },
                "modeBarButtonsToRemove": [
                    "lasso2d",
                    "select2d",
                ],
            },
            key=(
                f"plotly_order_"
                f"{order_label}_"
                f"{vin}"
            ),
        )

        st.caption(
            "Use the toolbar to zoom, pan, reset the axes or export the "
            "current graph as a PNG image. Click legend items to hide or "
            "restore individual curves."
        )
        return

    # -------------------------------------------------------------------------
    # Altair fallback
    # -------------------------------------------------------------------------

    chart_frames = []

    for channel_name, curve in valid_curves.items():
        chart_frames.append(
            pd.DataFrame(
                {
                    "RPM": curve["rpm"],
                    "Amplitude": curve["amp"],
                    "Series": channel_name,
                    "Series Type": "Measured",
                }
            )
        )

    if (
        target_rpm is not None
        and target_amp is not None
    ):
        target_rpm_values = np.asarray(
            target_rpm,
            dtype=float,
        )

        target_amp_values = np.asarray(
            target_amp,
            dtype=float,
        )

        target_mask = (
            np.isfinite(target_rpm_values)
            & np.isfinite(target_amp_values)
        )

        if np.any(target_mask):
            chart_frames.append(
                pd.DataFrame(
                    {
                        "RPM": target_rpm_values[
                            target_mask
                        ],
                        "Amplitude": target_amp_values[
                            target_mask
                        ],
                        "Series": "Target",
                        "Series Type": "Target",
                    }
                )
            )

    chart_data = pd.concat(
        chart_frames,
        ignore_index=True,
    )

    nearest = alt.selection_point(
        nearest=True,
        on="pointerover",
        fields=[
            "RPM"
        ],
        empty=False,
        clear="pointerout",
    )

    base = alt.Chart(
        chart_data
    ).encode(
        x=alt.X(
            "RPM:Q",
            title="Engine Speed [rpm]",
        ),
        y=alt.Y(
            "Amplitude:Q",
            title="Order Amplitude [m/s²]",
        ),
        color=alt.Color(
            "Series:N",
            scale=alt.Scale(
                domain=[
                    "ChA",
                    "ChB",
                    "ChC",
                    "Target",
                ],
                range=[
                    "#1768A6",
                    "#E67E22",
                    "#2E8B57",
                    "#C0392B",
                ],
            ),
        ),
        strokeDash=alt.StrokeDash(
            "Series Type:N",
            legend=None,
        ),
    )

    lines = base.mark_line(
        strokeWidth=2.4,
    )

    points = base.mark_circle(
        size=70,
    ).encode(
        opacity=alt.condition(
            nearest,
            alt.value(1),
            alt.value(0),
        ),
        tooltip=[
            "Series:N",
            alt.Tooltip(
                "RPM:Q",
                format=".0f",
            ),
            alt.Tooltip(
                "Amplitude:Q",
                format=".3f",
            ),
        ],
    ).add_params(
        nearest
    )

    fallback_chart = (
        alt.layer(
            lines,
            points,
        )
        .properties(
            height=520,
            title=(
                f"{order_label} — Interactive Order Response"
            ),
        )
        .interactive(
            bind_y=False,
        )
    )

    st.warning(
        "Plotly is not installed. The chart is being displayed with "
        "the Altair fallback. Add 'plotly' to requirements.txt to enable "
        "the full engineering toolbar and image export."
    )

    st.altair_chart(
        fallback_chart,
        width="stretch",
        theme=None,
    )

def render_interactive_order_map(
    time: np.ndarray,
    rpm: np.ndarray,
    signal: np.ndarray,
    selected_channel: str,
    analysis_type: str,
    vin: str,
    samples_per_rev: int,
    revs_per_block: int,
    overlap: float,
    max_order: float,
    calibration_factor: float,
) -> bool:
    """
    Render an interactive Plotly order map.

    Returns True when the Plotly chart is rendered and False when Plotly
    is unavailable, allowing the caller to use the static fallback.
    """
    if go is None:
        return False

    engine_angular_resample = (
        tc_angular_resample
        if analysis_type == ANALYSIS_TRANSFER_CASE
        else axle_angular_resample
    )

    engine_order_map = (
        tc_order_map
        if analysis_type == ANALYSIS_TRANSFER_CASE
        else axle_order_map
    )

    theta_u, signal_u, rpm_u = engine_angular_resample(
        time,
        rpm,
        signal,
        samples_per_rev=samples_per_rev,
    )

    orders, block_rpms, spectrum = engine_order_map(
        theta_u,
        signal_u,
        rpm_u,
        samples_per_rev=samples_per_rev,
        revs_per_block=revs_per_block,
        overlap=overlap,
        max_order=max_order,
    )

    orders = np.asarray(
        orders,
        dtype=float,
    )

    block_rpms = np.asarray(
        block_rpms,
        dtype=float,
    )

    spectrum = np.asarray(
        spectrum,
        dtype=float,
    )

    if spectrum.ndim != 2:
        raise ValueError(
            "Order spectrum must be two-dimensional."
        )

    if spectrum.shape != (
        len(block_rpms),
        len(orders),
    ):
        raise ValueError(
            "Order spectrum dimensions do not match the RPM and order axes."
        )

    sort_index = np.argsort(
        block_rpms,
        kind="stable",
    )

    sorted_rpm = block_rpms[
        sort_index
    ]

    sorted_spectrum = spectrum[
        sort_index,
        :
    ]

    decibels = 20.0 * np.log10(
        np.maximum(
            sorted_spectrum
            * calibration_factor,
            1e-12,
        )
    )

    figure = go.Figure(
        data=go.Heatmap(
            x=orders,
            y=sorted_rpm,
            z=decibels,
            colorscale="Turbo",
            colorbar={
                "title": {
                    "text": "dB re 1 m/s²",
                    "side": "right",
                },
                "tickfont": {
                    "color": "#536979",
                },
                "outlinecolor": "#CBD7DF",
                "outlinewidth": 1,
            },
            hovertemplate=(
                "<b>Order Map</b><br>"
                "Order: %{x:.2f}<br>"
                "RPM: %{y:.0f}<br>"
                "Amplitude: %{z:.2f} dB"
                "<extra></extra>"
            ),
            zsmooth=False,
        )
    )

    figure.update_layout(
        title={
            "text": (
                f"<b>Order Map / Waterfall — {selected_channel}</b>"
                f"<br><sup>VIN: {vin} &nbsp; | &nbsp; "
                f"{analysis_type} &nbsp; | &nbsp; "
                f"Max Order: {max_order:.0f}</sup>"
            ),
            "x": 0.01,
            "xanchor": "left",
            "font": {
                "size": 18,
                "color": "#17324D",
            },
        },
        height=640,
        margin={
            "l": 75,
            "r": 80,
            "t": 95,
            "b": 70,
        },
        paper_bgcolor="#F5F7FA",
        plot_bgcolor="#FFFFFF",
        xaxis={
            "title": {
                "text": "Order",
                "font": {
                    "color": "#30485C",
                    "size": 13,
                },
            },
            "showgrid": False,
            "linecolor": "#AEBCC7",
            "tickfont": {
                "color": "#536979",
            },
        },
        yaxis={
            "title": {
                "text": "Engine Speed [rpm]",
                "font": {
                    "color": "#30485C",
                    "size": 13,
                },
            },
            "showgrid": False,
            "linecolor": "#AEBCC7",
            "tickfont": {
                "color": "#536979",
            },
        },
        modebar={
            "bgcolor": "rgba(255,255,255,0.75)",
            "color": "#607585",
            "activecolor": "#1768A6",
        },
    )

    st.plotly_chart(
        figure,
        width="stretch",
        config={
            "displaylogo": False,
            "scrollZoom": True,
            "responsive": True,
            "toImageButtonOptions": {
                "format": "png",
                "filename": (
                    f"{vin}_{selected_channel}_order_map"
                ),
                "height": 720,
                "width": 1280,
                "scale": 2,
            },
            "modeBarButtonsToRemove": [
                "lasso2d",
                "select2d",
            ],
        },
        key=(
            f"plotly_order_map_"
            f"{selected_channel}_"
            f"{vin}"
        ),
    )

    st.caption(
        "Hover over the heatmap to inspect exact Order, RPM and dB values. "
        "Use the toolbar to zoom, pan, reset or export the current view."
    )

    return True


def display_order_result(
    order_value: float,
    definition: Mapping[str, object],
    result_dataframe: pd.DataFrame,
    channel_curves: Mapping[str, dict],
    result_vin: str,
    result_analysis_type: str,
    result_vehicle_configuration: str,
) -> None:
    status = determine_order_status(result_dataframe)

    section_title(f"{definition['label']} Result Summary")

    metric_columns = st.columns(4)

    for metric_column, channel_name in zip(
        metric_columns[:3],
        CHANNEL_NAMES,
    ):
        matching = result_dataframe.loc[
            result_dataframe["Channel"] == channel_name,
            "Peak Amplitude [m/s²]",
        ]
        value = (
            f"{float(matching.iloc[0]):.2f} m/s²"
            if len(matching)
            else "N/A"
        )
        metric_column.metric(
            f"Peak {channel_name}",
            value,
        )

    metric_columns[3].metric("Assessment", status)

    peak_rows = result_dataframe[
        result_dataframe["Peak Amplitude [m/s²]"].notna()
    ]

    if not peak_rows.empty:
        critical_row = peak_rows.loc[
            peak_rows["Peak Amplitude [m/s²]"].astype(float).idxmax()
        ]
        target_value = critical_row.get("Target at Peak RPM [m/s²]", np.nan)
        margin_value = critical_row.get("Max Margin [m/s²]", np.nan)

        card_columns = st.columns(4)
        card_values = [
            ("Critical Channel", str(critical_row.get("Channel","N/A")),
             "Highest measured response"),
            ("Peak Response",
             f"{float(critical_row['Peak Amplitude [m/s²]']):.2f} m/s²",
             f"at {float(critical_row.get('Peak RPM', np.nan)):.0f} rpm"),
            ("Target at Peak",
             f"{float(target_value):.2f} m/s²" if pd.notna(target_value) else "N/A",
             "Reference target level"),
            ("Maximum Margin",
             f"{float(margin_value):+.2f} m/s²" if pd.notna(margin_value) else "N/A",
             f"Assessment: {status}"),
        ]

        for column, (title, value, caption) in zip(card_columns, card_values):
            with column:
                st.markdown(
                    f"""
<div class="result-summary-card">
<div class="result-card-title">{title}</div>
<div class="result-card-value">{value}</div>
<div class="result-card-caption">{caption}</div>
</div>
""",
                    unsafe_allow_html=True,
                )

    render_interactive_order_chart(
        order_label=definition["label"],
        channel_curves=channel_curves,
        target_rpm=definition.get("target_rpm"),
        target_amp=definition.get("target_amp"),
        vin=result_vin,
        analysis_type=result_analysis_type,
        vehicle_configuration=result_vehicle_configuration,
    )

    if show_static_plots:
        with st.expander(
            "Static Engineering Plot",
            expanded=False,
        ):
            st.caption(
                "Static Matplotlib view used for report-style inspection."
            )

            figure = plot_order_comparison(
                order_label=definition["label"],
                channel_curves=channel_curves,
                target_rpm=definition.get("target_rpm"),
                target_amp=definition.get("target_amp"),
                vin=result_vin,
                analysis_type=result_analysis_type,
                vehicle_configuration=result_vehicle_configuration,
            )

            st.pyplot(
                figure,
                width="stretch",
            )

            plt.close(
                figure
            )

    st.dataframe(
        result_dataframe,
        width="stretch",
        hide_index=True,
    )

    if status == "PASS":
        st.success(f"{definition['label']} Assessment: PASS")
    elif status == "FAIL":
        st.error(f"{definition['label']} Assessment: FAIL")
    else:
        st.info(
            f"{definition['label']} Assessment: INFO — "
            "no target is defined for this harmonic."
        )


st.markdown('<div id="results-section"></div>', unsafe_allow_html=True)

if st.session_state.get("analysis_completed", False):
    st.divider()

    result_analysis_type = st.session_state["analysis_type_result"]
    result_time = st.session_state["time_result"]
    result_rpm = st.session_state["rpm_result"]
    result_channels = st.session_state["channels_result"]
    result_curves = st.session_state["curves_by_order"]
    result_tables = st.session_state["results_by_order"]
    result_raw_curves = st.session_state["raw_curves_by_order"]
    result_definitions = st.session_state["order_definitions_result"]
    result_status = st.session_state["overall_status"]
    result_configuration = st.session_state[
        "vehicle_configuration_result"
    ]
    result_channel = st.session_state["selected_channel_result"]
    result_settings = st.session_state["analysis_settings_result"]
    result_vin = st.session_state["vin_result"]

    render_analysis_dashboard(
        result_tables=result_tables,
        overall_status=result_status,
        vin=result_vin,
        analysis_type=result_analysis_type,
        vehicle_configuration=result_configuration,
    )

    with st.container(border=True):
        section_title(
            "Analysis Report",
            "Download complete calculation and compliance reports.",
        )

        report_columns = st.columns(2)

        with report_columns[0]:
            st.download_button(
                label="Download Excel Report",
                data=st.session_state["excel_report"],
                file_name=(
                    f"{result_vin}_"
                    f"{result_analysis_type.replace(' ', '_')}"
                    "_report.xlsx"
                ),
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                width="stretch",
                key="download_excel_report",
            )

        with report_columns[1]:
            pdf_report = st.session_state.get("pdf_report")

            if pdf_report is not None:
                st.download_button(
                    label=T["pdf"],
                    data=pdf_report,
                    file_name=(
                        f"{result_vin}_"
                        f"{result_analysis_type.replace(' ', '_')}"
                        "_report.pdf"
                    ),
                    mime="application/pdf",
                    width="stretch",
                    key="download_pdf_report",
                )
            else:
                st.button(
                    T["pdf"],
                    width="stretch",
                    disabled=True,
                    help="Add reportlab to requirements.txt to enable PDF reporting.",
                    key="pdf_report_unavailable",
                )

    if result_analysis_type == ANALYSIS_TRANSFER_CASE:
        order_results_tab, order_map_tab, raw_tab = st.tabs(
            [
                "Gear Mesh Order Results",
                "Order Map / Waterfall",
                "Raw Results",
            ]
        )

        with order_results_tab:
            for order_value in [63.0, 85.05, 126.0, 170.10]:
                if (
                    order_value in result_definitions
                    and order_value in result_tables
                ):
                    with st.container(border=True):
                        display_order_result(
                            order_value,
                            result_definitions[order_value],
                            result_tables[order_value],
                            result_curves[order_value],
                            result_vin,
                            result_analysis_type,
                            result_configuration,
                        )
    else:
        order_10_tab, order_20_tab, order_map_tab, raw_tab = st.tabs(
            [
                "10th Order",
                "20th Order",
                "Order Map / Waterfall",
                "Raw Results",
            ]
        )

        for order_value, tab in [
            (10.0, order_10_tab),
            (20.0, order_20_tab),
        ]:
            with tab:
                with st.container(border=True):
                    display_order_result(
                        order_value,
                        result_definitions[order_value],
                        result_tables[order_value],
                        result_curves[order_value],
                        result_vin,
                        result_analysis_type,
                        result_configuration,
                    )

    with order_map_tab:
        with st.container(border=True):
            section_title(
                f"Order Map / Waterfall — {result_channel}"
            )
            interactive_map_rendered = render_interactive_order_map(
                time=result_time,
                rpm=result_rpm,
                signal=result_channels[
                    result_channel
                ],
                selected_channel=result_channel,
                analysis_type=result_analysis_type,
                vin=result_vin,
                samples_per_rev=result_settings[
                    "samples_per_rev"
                ],
                revs_per_block=result_settings[
                    "revs_per_block"
                ],
                overlap=result_settings[
                    "overlap"
                ],
                max_order=result_settings[
                    "max_order"
                ],
                calibration_factor=result_settings[
                    "calibration_factor"
                ],
            )

            if (
                show_static_plots
                or not interactive_map_rendered
            ):
                with st.expander(
                    "Static Engineering Waterfall",
                    expanded=not interactive_map_rendered,
                ):
                    if not interactive_map_rendered:
                        st.warning(
                            "Plotly is not installed. Displaying the static "
                            "Matplotlib waterfall."
                        )

                    map_figure = create_order_map_figure(
                        time=result_time,
                        rpm=result_rpm,
                        signal=result_channels[
                            result_channel
                        ],
                        selected_channel=result_channel,
                        analysis_type=result_analysis_type,
                        vin=result_vin,
                        samples_per_rev=result_settings[
                            "samples_per_rev"
                        ],
                        revs_per_block=result_settings[
                            "revs_per_block"
                        ],
                        overlap=result_settings[
                            "overlap"
                        ],
                        max_order=result_settings[
                            "max_order"
                        ],
                        calibration_factor=result_settings[
                            "calibration_factor"
                        ],
                    )

                    st.pyplot(
                        map_figure,
                        width="stretch",
                    )

                    plt.close(
                        map_figure
                    )

    with raw_tab:
        for order_value, curve_dataframe in result_raw_curves.items():
            with st.container(border=True):
                section_title(
                    f"{result_definitions[order_value]['label']} Raw Curve Data"
                )
                st.dataframe(
                    curve_dataframe,
                    width="stretch",
                    hide_index=True,
                )


st.markdown('<div id="system-section"></div>', unsafe_allow_html=True)

# =============================================================================
# SYSTEM STATUS
# =============================================================================

with st.expander("System and Module Status", expanded=False):
    st.success("Axle Whine analysis engine: Ready")
    st.success("Transfer Case analysis engine: Ready")
    st.write(
        {
            "Supported file types": ["XLSX", "CSV"],
            "Supported channels": CHANNEL_NAMES,
            "CSV unit conversion": "g → m/s²",
        }
    )
